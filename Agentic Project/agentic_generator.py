# crewai_test_case_generator.py

"""
CrewAI-powered Automated Test Case Generation System
Converts your existing RAG system into an agentic workflow
"""

import os
import json
import requests
import numpy as np
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from langchain_huggingface import HuggingFaceEmbeddings

# CrewAI imports
from crewai import Agent, Task, Crew
from crewai.tools import BaseTool
from langchain_community.chat_models import ChatOllama

# ----------------------------
# EXISTING SETUP (unchanged)
# ----------------------------
load_dotenv()

# Database connection
DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASSWORD")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_SCHEMA = os.getenv("DB_SCHEMA", "qpot")

DATABASE_URL = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(DATABASE_URL)

# AI settings
OLLAMA_URL = os.getenv("OLLAMA_URL")
MODEL_NAME = os.getenv("MODEL_NAME", "mistral:7b")
embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")

OUTPUT_DIR = "test_cases"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# CrewAI LLM configuration
local_llm = ChatOllama(
    model=MODEL_NAME,
    base_url=OLLAMA_URL,
    temperature=0.1
)

# ----------------------------
# EXISTING FUNCTIONS (unchanged)
# ----------------------------

def find_excel_chunks(question, top_k=8):
    """Find chunks specifically from Excel files (test case templates)"""
    try:
        question_vector = np.array(embeddings.embed_query(question))
        
        with engine.connect() as conn:
            conn.execute(text(f"SET search_path TO {DB_SCHEMA}, public"))
            result = conn.execute(text("""
                SELECT doc_name, chunk_text, embedding, file_type 
                FROM vector_documents 
                WHERE embedding IS NOT NULL 
                AND file_type = 'excel'
                ORDER BY id
            """))
            
            similar_chunks = []
            for row in result:
                doc_name, chunk_text, embedding, file_type = row
                
                try:
                    if isinstance(embedding, str):
                        embedding_str = embedding.strip('[]')
                        embedding_list = [float(x.strip()) for x in embedding_str.split(',')]
                        embedding_vector = np.array(embedding_list)
                    else:
                        embedding_vector = np.array(embedding)
                    
                    dot_product = np.dot(question_vector, embedding_vector)
                    norm1 = np.linalg.norm(question_vector)
                    norm2 = np.linalg.norm(embedding_vector)
                    
                    if norm1 > 0 and norm2 > 0:
                        similarity = dot_product / (norm1 * norm2)
                        similar_chunks.append((similarity, doc_name, chunk_text, file_type))
                
                except Exception:
                    continue
            
            similar_chunks.sort(reverse=True)
            return similar_chunks[:top_k]
    
    except Exception as e:
        print(f"Error finding Excel chunks: {e}")
        return []

def find_workflow_chunks(question, top_k=3):
    """Find chunks from Word documents for workflow understanding"""
    try:
        question_vector = np.array(embeddings.embed_query(question))
        
        with engine.connect() as conn:
            conn.execute(text(f"SET search_path TO {DB_SCHEMA}, public"))
            result = conn.execute(text("""
                SELECT doc_name, chunk_text, embedding, file_type 
                FROM vector_documents 
                WHERE embedding IS NOT NULL 
                AND file_type = 'word'
                ORDER BY id
            """))
            
            similar_chunks = []
            for row in result:
                doc_name, chunk_text, embedding, file_type = row
                
                try:
                    if isinstance(embedding, str):
                        embedding_str = embedding.strip('[]')
                        embedding_list = [float(x.strip()) for x in embedding_str.split(',')]
                        embedding_vector = np.array(embedding_list)
                    else:
                        embedding_vector = np.array(embedding)
                    
                    dot_product = np.dot(question_vector, embedding_vector)
                    norm1 = np.linalg.norm(question_vector)
                    norm2 = np.linalg.norm(embedding_vector)
                    
                    if norm1 > 0 and norm2 > 0:
                        similarity = dot_product / (norm1 * norm2)
                        similar_chunks.append((similarity, doc_name, chunk_text, file_type))
                
                except Exception:
                    continue
            
            similar_chunks.sort(reverse=True)
            return similar_chunks[:top_k]
    
    except Exception as e:
        print(f"Error finding workflow chunks: {e}")
        return []

def find_similar_chunks_general(question):
    """General purpose chunk search"""
    try:
        question_vector = np.array(embeddings.embed_query(question))
        
        with engine.connect() as conn:
            conn.execute(text(f"SET search_path TO {DB_SCHEMA}, public"))
            result = conn.execute(text("SELECT doc_name, chunk_text, embedding FROM vector_documents WHERE embedding IS NOT NULL"))
            
            similar_chunks = []
            for row in result:
                doc_name, chunk_text, embedding = row
                
                try:
                    if isinstance(embedding, str):
                        embedding_str = embedding.strip('[]')
                        embedding_list = [float(x.strip()) for x in embedding_str.split(',')]
                        embedding_vector = np.array(embedding_list)
                    elif isinstance(embedding, list):
                        embedding_vector = np.array(embedding)
                    else:
                        embedding_vector = np.array(embedding)
                    
                    dot_product = np.dot(question_vector, embedding_vector)
                    norm1 = np.linalg.norm(question_vector)
                    norm2 = np.linalg.norm(embedding_vector)
                    
                    if norm1 > 0 and norm2 > 0:
                        similarity = dot_product / (norm1 * norm2)
                        similar_chunks.append((similarity, doc_name, chunk_text))
                
                except Exception as calc_error:
                    continue
            
            similar_chunks.sort(reverse=True)
            return similar_chunks[:3]
    
    except Exception as e:
        print(f"Error finding chunks: {e}")
        return []

def create_excel_from_response(ai_response, filename_prefix="test_cases"):
    """Convert AI response to Excel file with colorful formatting"""
    try:
        start_idx = ai_response.find('{')
        end_idx = ai_response.rfind('}') + 1
        
        if start_idx == -1 or end_idx == 0:
            return create_simple_excel(ai_response, filename_prefix)
        
        json_str = ai_response[start_idx:end_idx]
        test_data = json.loads(json_str)
        
        excel_data = []
        
        for scenario in test_data.get('scenarios', []):
            scenario_id = scenario.get('scenario_id', '')
            scenario_desc = scenario.get('scenario_description', '')
            area = scenario.get('area', '')
            module_name = scenario.get('module_name', '')
            sub_module_name = scenario.get('sub_module_name', '')
            sub_sub_module_name = scenario.get('sub_sub_module_name', '')
            
            for test_case in scenario.get('test_cases', []):
                excel_data.append({
                    'TestCaseID': test_case.get('test_case_id', ''),
                    'TestScenarioID': scenario_id,
                    'Area': area,
                    'ModuleName': module_name,
                    'SubModuleName': sub_module_name,
                    'TestScenarioDescription': scenario_desc,
                    'SubSubModuleName': sub_sub_module_name,
                    'Priority': test_case.get('priority', ''),
                    'TestCaseDescription': test_case.get('test_case_description', ''),
                    'PreRequisites': test_case.get('prerequisites', ''),
                    'TestData': test_case.get('test_data', ''),
                    'NavigationSteps': test_case.get('navigation_steps', ''),
                    'ExpectedOutput': test_case.get('expected_output', ''),
                    'ActualOutput': test_case.get('actual_output', ''),
                    'Status': test_case.get('status', 'Not Executed')
                })
        
        # Create formatted Excel (your existing logic)
        df = pd.DataFrame(excel_data)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Generated_Test_Cases"
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply your existing formatting logic here...
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(filepath)
        return filepath, len(excel_data)
    
    except Exception as e:
        return create_simple_excel(ai_response, filename_prefix)

def create_simple_excel(ai_response, filename_prefix):
    """Create simple Excel file when JSON parsing fails"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}_raw.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    df = pd.DataFrame({
        'Generated_Content': [ai_response],
        'Generation_Time': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        'Note': ['Please review and format manually']
    })
    
    df.to_excel(filepath, sheet_name='Raw_Output', index=False)
    return filepath, 1

# ----------------------------
# CREWAI TOOLS (New - wrapping existing functions)
# ----------------------------

class RAGSearchTool(BaseTool):
    name: str = "RAG Document Search"
    description: str = "Search through vector documents for relevant testing information"
    
    def _run(self, query: str) -> str:
        chunks = find_similar_chunks_general(query)
        context = "Relevant document information:\n\n"
        for similarity, doc_name, chunk_text in chunks:
            context += f"From {doc_name}:\n{chunk_text}\n\n"
        return context

class ExcelTemplateTool(BaseTool):
    name: str = "Excel Template Analyzer" 
    description: str = "Find Excel test case templates and structures"
    
    def _run(self, request_type: str) -> str:
        excel_chunks = find_excel_chunks(request_type, top_k=10)
        template_info = "Test case template structures:\n\n"
        for similarity, doc_name, chunk_text, file_type in excel_chunks:
            template_info += f"Template from {doc_name}:\n{chunk_text}\n\n"
        return template_info

class WorkflowAnalysisTool(BaseTool):
    name: str = "Workflow Documentation Analyzer"
    description: str = "Analyze workflow documentation for navigation steps"
    
    def _run(self, functionality: str) -> str:
        workflow_chunks = find_workflow_chunks(functionality, top_k=5)
        workflow_info = "Workflow and navigation information:\n\n"
        for similarity, doc_name, chunk_text, file_type in workflow_chunks:
            workflow_info += f"Workflow from {doc_name}:\n{chunk_text}\n\n"
        return workflow_info

class TestCaseGeneratorTool(BaseTool):
    name: str = "Test Case JSON Generator"
    description: str = "Generate test cases in JSON format using Ollama"
    
    def _run(self, context: str) -> str:
        prompt = f"""Based on the following context, generate EXACTLY 5 test scenarios with 5 test cases each (25 total).

{context}

Return ONLY valid JSON in this exact format:
{{
    "scenarios": [
        {{
            "scenario_id": "TS001",
            "scenario_description": "Brief description of scenario",
            "area": "Testing area (e.g., Login, Payment)",
            "module_name": "Module name",
            "sub_module_name": "Sub module (if any)",
            "sub_sub_module_name": "Sub sub module (if any)",
            "test_cases": [
                {{
                    "test_case_id": "TC001",
                    "priority": "High/Medium/Low", 
                    "test_case_description": "Test case description",
                    "prerequisites": "Prerequisites",
                    "test_data": "Test data",
                    "navigation_steps": "Navigation steps to reproduce (use -> for steps)",
                    "expected_output": "Expected output",
                    "actual_output": "",
                    "status": "Not Executed"
                }}
            ]
        }}
    ]
}}"""
        
        response = requests.post(
            f"{OLLAMA_URL}/api/generate",
            json={"model": MODEL_NAME, "prompt": prompt, "stream": False},
            timeout=120
        )
        
        if response.status_code == 200:
            return response.json().get('response', '')
        else:
            return f"Error: {response.status_code}"

# ----------------------------
# CREWAI AGENTS
# ----------------------------

# Requirements Analyst Agent
requirements_analyst = Agent(
    role='Test Requirements Analyst',
    goal='Analyze testing requirements and determine comprehensive test scope',
    backstory="""You are an expert test analyst with 10+ years of experience in software testing. 
    You excel at breaking down complex testing requirements and identifying all necessary test scenarios.""",
    tools=[RAGSearchTool()],
    llm=local_llm,
    verbose=True,
    allow_delegation=False
)

# Test Designer Agent
test_designer = Agent(
    role='Test Case Designer',
    goal='Design detailed test scenarios based on requirements analysis',
    backstory="""You are a senior test designer who creates comprehensive test scenarios.
    You understand test case templates and can design tests that cover all edge cases and workflows.""",
    tools=[ExcelTemplateTool(), WorkflowAnalysisTool()],
    llm=local_llm,
    verbose=True,
    allow_delegation=False
)

# Test Generator Agent
test_generator = Agent(
    role='Test Case Generator',
    goal='Generate actual executable test cases in proper JSON format',
    backstory="""You are a test automation specialist who generates precise test cases.
    You follow templates exactly and ensure all test cases are complete and executable.""",
    tools=[TestCaseGeneratorTool()],
    llm=local_llm,
    verbose=True,
    allow_delegation=False
)

# ----------------------------
# CREWAI TASKS
# ----------------------------

def create_analysis_task(testing_request):
    return Task(
        description=f"""Analyze this testing request comprehensively: {testing_request}
        
        Your analysis should include:
        1. Type of testing required (functional, integration, etc.)
        2. Key areas and modules that need testing
        3. Priority levels for different test scenarios
        4. Relevant documentation and templates to use
        
        Provide a detailed analysis that will guide the test design process.""",
        agent=requirements_analyst,
        expected_output="Comprehensive testing requirements analysis with priorities and scope"
    )

def create_design_task():
    return Task(
        description="""Based on the requirements analysis, design comprehensive test scenarios.
        
        Your design should include:
        1. 5 distinct test scenarios covering different aspects
        2. Test case templates and structures to follow  
        3. Workflow navigation requirements
        4. Test data and prerequisite requirements
        5. Expected coverage areas for each scenario
        
        Focus on creating designs that will result in thorough test coverage.""",
        agent=test_designer,
        expected_output="Detailed test scenario designs with templates and workflow information"
    )

def create_generation_task():
    return Task(
        description="""Generate the actual test cases based on the designed scenarios.
        
        Requirements:
        1. Generate exactly 5 test scenarios
        2. Each scenario must have exactly 5 test cases (25 total)
        3. Follow the Excel template structure exactly
        4. Include all required fields for each test case
        5. Return in valid JSON format only
        
        The output must be ready for direct Excel conversion.""",
        agent=test_generator,
        expected_output="Complete test cases in valid JSON format ready for Excel export"
    )

# ----------------------------
# CREWAI ORCHESTRATION
# ----------------------------

def process_testing_request_with_crew(testing_request: str):
    """Process testing request using CrewAI agents"""
    print(f"🚀 CrewAI Processing: {testing_request}")
    
    # Create tasks
    analysis_task = create_analysis_task(testing_request)
    design_task = create_design_task()
    generation_task = create_generation_task()
    
    # Create crew
    test_crew = Crew(
        agents=[requirements_analyst, test_designer, test_generator],
        tasks=[analysis_task, design_task, generation_task],
        verbose=2,
        process="sequential"
    )
    
    try:
        # Execute crew
        print("🤖 Crew executing tasks...")
        result = test_crew.kickoff()
        
        # Convert result to Excel
        print("📊 Converting to Excel...")
        filepath, count = create_excel_from_response(str(result), "crewai_test_cases")
        
        return {
            "success": True,
            "filepath": filepath,
            "count": count,
            "crew_output": str(result)
        }
        
    except Exception as e:
        print(f"❌ CrewAI execution failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }

# ----------------------------
# AUTO-TRIGGER SYSTEM
# ----------------------------

def monitor_requests_folder():
    """Monitor requests folder for automatic processing"""
    requests_folder = "requests"
    os.makedirs(requests_folder, exist_ok=True)
    
    print(f"👀 Monitoring {requests_folder}/ for new requests...")
    print("📝 Create a JSON file with: {\"request\": \"Generate test cases for login module\"}")
    
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    
    class RequestProcessor(FileSystemEventHandler):
        def on_created(self, event):
            if event.is_file and event.src_path.endswith('.json'):
                print(f"📥 New request detected: {event.src_path}")
                
                try:
                    with open(event.src_path, 'r') as f:
                        request_data = json.load(f)
                    
                    testing_request = request_data.get('request', '')
                    if testing_request:
                        result = process_testing_request_with_crew(testing_request)
                        
                        if result['success']:
                            print(f"✅ Generated: {result['filepath']} ({result['count']} test cases)")
                        else:
                            print(f"❌ Error: {result['error']}")
                    
                    # Clean up request file
                    os.remove(event.src_path)
                    
                except Exception as e:
                    print(f"❌ Error processing request: {e}")
    
    # Start monitoring
    event_handler = RequestProcessor()
    observer = Observer()
    observer.schedule(event_handler, requests_folder, recursive=False)
    observer.start()
    
    try:
        import time
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("👋 Stopping request monitor...")
    
    observer.join()

# ----------------------------
# MAIN INTERFACE
# ----------------------------

def main():
    print("🤖 CREWAI AUTOMATED TEST CASE GENERATION SYSTEM")
    print("=" * 60)
    print("AGENTIC WORKFLOW: Multi-agent collaborative test generation")
    print("=" * 60)
    
    while True:
        print("\nCHOOSE MODE:")
        print("1. 🧪 Manual Test Case Generation")
        print("2. 👀 Auto-Monitor Requests Folder") 
        print("3. 🚪 Quit")
        
        choice = input("\nEnter choice (1/2/3): ").strip()
        
        if choice == '3' or choice.lower() in ['quit', 'exit']:
            print("👋 Goodbye!")
            break
            
        elif choice == '1':
            testing_request = input("\n🎯 Describe what test cases you need: ").strip()
            
            if testing_request:
                result = process_testing_request_with_crew(testing_request)
                
                if result['success']:
                    print(f"\n✅ SUCCESS!")
                    print(f"📁 File: {os.path.basename(result['filepath'])}")
                    print(f"📊 Test cases: {result['count']}")
                    print(f"🗂️ Location: test_cases/")
                else:
                    print(f"\n❌ Error: {result['error']}")
        
        elif choice == '2':
            monitor_requests_folder()
        
        else:
            print("❌ Invalid choice. Please enter 1, 2, or 3.")

if __name__ == "__main__":
    main()
