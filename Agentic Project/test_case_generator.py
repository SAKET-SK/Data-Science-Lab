# test_case_generator.py

"""
Automated Test Case Generation System using RAG
Specialized for generating test scenarios and test cases based on Excel templates.
Uses Word documents for workflow understanding and navigation steps.

STRICT USE CASE: Only for test case generation - refuses other queries.
"""

import os
import json
import requests
import numpy as np
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from langchain_huggingface import HuggingFaceEmbeddings

# ----------------------------
# SETUP
# ----------------------------
load_dotenv()

# Database connection
DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASSWORD")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_SCHEMA = os.getenv("DB_SCHEMA", "qpot")

DATABASE_URL = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(DATABASE_URL)

# AI settings
OLLAMA_URL = os.getenv("OLLAMA_URL")
MODEL_NAME = os.getenv("MODEL_NAME", "mistral:7b")
embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")

# Output directory for generated test cases
OUTPUT_DIR = "test_cases"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ----------------------------
# VALIDATION FUNCTIONS
# ----------------------------

def is_test_case_request(question):
    """Check if the question is related to test case generation"""
    test_keywords = [
        'test case', 'test scenario', 'testing', 'scenario', 'test',
        'generate', 'create', 'automation', 'validation', 'verify',
        'functionality', 'feature', 'module', 'workflow', 'steps'
    ]
    
    question_lower = question.lower()
    return any(keyword in question_lower for keyword in test_keywords)

def refuse_non_test_queries(question):
    """Refuse to answer non-test case related questions"""
    return """
🚫 ACCESS DENIED

This system is STRICTLY designed for Automated Test Case Generation only.

✅ Valid requests:
- Generate test cases for [feature/module]
- Create test scenarios for [functionality] 
- Generate automated test cases
- Create testing scenarios
- Test case generation for [specific area]

❌ Invalid requests:
- General questions
- Non-testing related queries
- Information lookup
- Documentation questions

Please ask only about TEST CASE GENERATION.
"""

# ----------------------------
# SEARCH FUNCTIONS
# ----------------------------

def find_excel_chunks(question, top_k=8):
    """Find chunks specifically from Excel files (test case templates)"""
    try:
        question_vector = np.array(embeddings.embed_query(question))
        
        with engine.connect() as conn:
            conn.execute(text(f"SET search_path TO {DB_SCHEMA}, public"))
            # Prioritize Excel files for test case structure
            result = conn.execute(text("""
                SELECT doc_name, chunk_text, embedding, file_type 
                FROM vector_documents 
                WHERE embedding IS NOT NULL 
                AND file_type = 'excel'
                ORDER BY id
            """))
            
            similar_chunks = []
            for row in result:
                doc_name, chunk_text, embedding, file_type = row
                
                try:
                    if isinstance(embedding, str):
                        embedding_str = embedding.strip('[]')
                        embedding_list = [float(x.strip()) for x in embedding_str.split(',')]
                        embedding_vector = np.array(embedding_list)
                    else:
                        embedding_vector = np.array(embedding)
                    
                    dot_product = np.dot(question_vector, embedding_vector)
                    norm1 = np.linalg.norm(question_vector)
                    norm2 = np.linalg.norm(embedding_vector)
                    
                    if norm1 > 0 and norm2 > 0:
                        similarity = dot_product / (norm1 * norm2)
                        similar_chunks.append((similarity, doc_name, chunk_text, file_type))
                
                except Exception:
                    continue
            
            similar_chunks.sort(reverse=True)
            return similar_chunks[:top_k]
    
    except Exception as e:
        print(f"Error finding Excel chunks: {e}")
        return []

def find_workflow_chunks(question, top_k=3):
    """Find chunks from Word documents for workflow understanding"""
    try:
        question_vector = np.array(embeddings.embed_query(question))
        
        with engine.connect() as conn:
            conn.execute(text(f"SET search_path TO {DB_SCHEMA}, public"))
            result = conn.execute(text("""
                SELECT doc_name, chunk_text, embedding, file_type 
                FROM vector_documents 
                WHERE embedding IS NOT NULL 
                AND file_type = 'word'
                ORDER BY id
            """))
            
            similar_chunks = []
            for row in result:
                doc_name, chunk_text, embedding, file_type = row
                
                try:
                    if isinstance(embedding, str):
                        embedding_str = embedding.strip('[]')
                        embedding_list = [float(x.strip()) for x in embedding_str.split(',')]
                        embedding_vector = np.array(embedding_list)
                    else:
                        embedding_vector = np.array(embedding)
                    
                    dot_product = np.dot(question_vector, embedding_vector)
                    norm1 = np.linalg.norm(question_vector)
                    norm2 = np.linalg.norm(embedding_vector)
                    
                    if norm1 > 0 and norm2 > 0:
                        similarity = dot_product / (norm1 * norm2)
                        similar_chunks.append((similarity, doc_name, chunk_text, file_type))
                
                except Exception:
                    continue
            
            similar_chunks.sort(reverse=True)
            return similar_chunks[:top_k]
    
    except Exception as e:
        print(f"Error finding workflow chunks: {e}")
        return []

# ----------------------------
# TEST CASE GENERATION
# ----------------------------

def generate_test_cases_with_rag(user_request):
    """Generate test cases using RAG service response and Excel templates"""
    print("🔍 Step 1: Getting RAG response for context...")
    
    # First, get general RAG response about the request
    rag_response = general_rag_answer(user_request)
    print(f"📋 RAG Context Retrieved: {len(rag_response)} characters")
    
    print("🔍 Step 2: Analyzing Excel test case templates...")
    # Get test case examples from Excel
    excel_chunks = find_excel_chunks(user_request, top_k=10)
    
    print("📋 Step 3: Understanding workflow from documentation...")
    # Get workflow information from Word docs
    workflow_chunks = find_workflow_chunks(user_request, top_k=5)
    
    if not excel_chunks:
        return "No test case templates found. Please ensure Excel files are processed."
    
    # Build enhanced context using RAG response + templates
    rag_context = f"RAG SERVICE CONTEXT:\n{rag_response}\n\n"
    
    excel_context = "TEST CASE EXAMPLES AND STRUCTURE:\n\n"
    for similarity, doc_name, chunk_text, file_type in excel_chunks:
        excel_context += f"From {doc_name}:\n{chunk_text}\n\n"
    
    workflow_context = "WORKFLOW AND NAVIGATION INFORMATION:\n\n"
    for similarity, doc_name, chunk_text, file_type in workflow_chunks:
        workflow_context += f"From {doc_name}:\n{chunk_text}\n\n"
    
    # Create enhanced prompt combining RAG response with templates
    prompt = f"""You are a specialized test case generation system. Use the RAG service context along with test case templates and workflow information to generate EXACTLY 5 test scenarios with 5 test cases each (25 total test cases).

{rag_context}

{excel_context}

{workflow_context}

User Request: {user_request}

INSTRUCTIONS:
1. Use the RAG service context to understand the functionality deeply
2. Generate EXACTLY 5 test scenarios based on this understanding
3. Each scenario must have EXACTLY 5 test cases
4. Follow the exact structure from the Excel templates
5. Use workflow information for detailed navigation steps
6. Make test cases realistic and executable based on RAG context
7. Include proper test data and expected outputs
8. Ensure comprehensive coverage of the requested functionality

FORMAT: Return as JSON with this exact structure:
{{
    "scenarios": [
        {{
            "scenario_id": "TS001",
            "scenario_description": "Description based on RAG context",
            "area": "Area name",
            "module_name": "Module name", 
            "sub_module_name": "Sub module name",
            "sub_sub_module_name": "Sub sub module name",
            "test_cases": [
                {{
                    "test_case_id": "TC001", 
                    "priority": "High/Medium/Low",
                    "test_case_description": "Description based on RAG understanding",
                    "prerequisites": "Prerequisites from context", 
                    "test_data": "Relevant test data",
                    "navigation_steps": "Detailed steps from workflow docs",
                    "expected_output": "Expected result based on functionality",
                    "actual_output": "",
                    "status": "Not Executed"
                }}
                // ... 4 more test cases
            ]
        }}
        // ... 4 more scenarios
    ]
}}

Generate the test cases now using the comprehensive context provided:"""

    print("🤖 Step 4: Generating comprehensive test cases with RAG context...")
    
    try:
        response = requests.post(
            f"{OLLAMA_URL}/api/generate",
            json={"model": MODEL_NAME, "prompt": prompt, "stream": False},
            timeout=120  # Longer timeout for complex generation
        )
        
        if response.status_code == 200:
            ai_response = response.json().get('response', '')
            print("✅ Test cases generated successfully using RAG context!")
            return ai_response
        else:
            return f"AI Error: {response.status_code}"
    
    except Exception as e:
        return f"Error generating test cases: {e}"

# ----------------------------
# EXCEL EXPORT FUNCTIONS
# ----------------------------

def create_excel_from_response(ai_response, filename_prefix="test_cases"):
    """Convert AI response to Excel file with colorful formatting matching the template"""
    try:
        # Try to extract JSON from AI response
        start_idx = ai_response.find('{')
        end_idx = ai_response.rfind('}') + 1
        
        if start_idx == -1 or end_idx == 0:
            # If no JSON found, create a simple text export
            return create_simple_excel(ai_response, filename_prefix)
        
        json_str = ai_response[start_idx:end_idx]
        test_data = json.loads(json_str)
        
        # Prepare data for Excel with exact column structure
        excel_data = []
        
        for scenario in test_data.get('scenarios', []):
            scenario_id = scenario.get('scenario_id', '')
            scenario_desc = scenario.get('scenario_description', '')
            area = scenario.get('area', '')
            module_name = scenario.get('module_name', '')
            sub_module_name = scenario.get('sub_module_name', '')
            sub_sub_module_name = scenario.get('sub_sub_module_name', '')
            
            for test_case in scenario.get('test_cases', []):
                excel_data.append({
                    'TestCaseID': test_case.get('test_case_id', ''),
                    'TestScenarioID': scenario_id,
                    'Area': area,
                    'ModuleName': module_name,
                    'SubModuleName': sub_module_name,
                    'TestScenarioDescription': scenario_desc,
                    'SubSubModuleName': sub_sub_module_name,
                    'Priority': test_case.get('priority', ''),
                    'TestCaseDescription': test_case.get('test_case_description', ''),
                    'PreRequisites': test_case.get('prerequisites', ''),
                    'TestData': test_case.get('test_data', ''),
                    'NavigationSteps': test_case.get('navigation_steps', ''),
                    'ExpectedOutput': test_case.get('expected_output', ''),
                    'ActualOutput': test_case.get('actual_output', ''),
                    'Status': test_case.get('status', 'Not Executed')
                })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        # Save with beautiful formatting like your template
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Generated_Test_Cases"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Define colors and styles like your template
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue header
        header_font = Font(color="FFFFFF", bold=True, size=10)  # White bold text
        
        # Priority color mapping
        priority_colors = {
            'High': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),    # Light Red
            'Medium': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),  # Light Yellow  
            'Low': PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")     # Light Green
        }
        
        # Row alternating colors (light blue)
        row_fill_1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Light blue
        row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Apply data formatting
        for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
            # Alternate row colors
            if row_num % 2 == 0:
                row_fill = row_fill_1
            else:
                row_fill = row_fill_2
            
            for col_num, cell in enumerate(row, 1):
                # Apply border and alignment
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Apply priority color if it's the Priority column (column H = 8)
                if col_num == 8 and cell.value in priority_colors:  # Priority column
                    cell.fill = priority_colors[cell.value]
                    cell.font = Font(bold=True)
                else:
                    cell.fill = row_fill
        
        # Set column widths to properly fit content (increased widths)
        column_widths = {
            'A': 15,  # TestCaseID
            'B': 18,  # TestScenarioID  
            'C': 20,  # Area
            'D': 25,  # ModuleName
            'E': 25,  # SubModuleName
            'F': 50,  # TestScenarioDescription
            'G': 25,  # SubSubModuleName
            'H': 15,  # Priority
            'I': 60,  # TestCaseDescription
            'J': 35,  # PreRequisites
            'K': 30,  # TestData
            'L': 60,  # NavigationSteps
            'M': 40,  # ExpectedOutput
            'N': 20,  # ActualOutput
            'O': 18   # Status
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Auto-adjust column widths based on content (with minimum widths set above)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Use the larger of: predefined width or content-based width
            predefined_width = column_widths.get(column_letter, 15)
            adjusted_width = max(predefined_width, min(max_length + 3, 80))  # Max 80 chars
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for better readability (increased heights)
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 35  # Increased from 25
        
        # Header row height
        ws.row_dimensions[1].height = 40  # Increased from 30
        
        # Save the workbook
        wb.save(filepath)
        
        return filepath, len(excel_data)
    
    except Exception as e:
        print(f"Error creating formatted Excel: {e}")
        return create_simple_excel(ai_response, filename_prefix)

def create_simple_excel(ai_response, filename_prefix):
    """Create simple Excel file when JSON parsing fails"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}_raw.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    # Create simple structure
    df = pd.DataFrame({
        'Generated_Content': [ai_response],
        'Generation_Time': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        'Note': ['Please review and format manually']
    })
    
    df.to_excel(filepath, sheet_name='Raw_Output', index=False)
    return filepath, 1

# ----------------------------
# GENERAL RAG FUNCTIONS (from rag_system.py)
# ----------------------------

def find_similar_chunks_general(question):
    """General purpose chunk search (from working rag_system.py)"""
    try:
        question_vector = np.array(embeddings.embed_query(question))
        
        with engine.connect() as conn:
            conn.execute(text(f"SET search_path TO {DB_SCHEMA}, public"))
            result = conn.execute(text("SELECT doc_name, chunk_text, embedding FROM vector_documents WHERE embedding IS NOT NULL"))
            
            similar_chunks = []
            for row in result:
                doc_name, chunk_text, embedding = row
                
                try:
                    if isinstance(embedding, str):
                        embedding_str = embedding.strip('[]')
                        embedding_list = [float(x.strip()) for x in embedding_str.split(',')]
                        embedding_vector = np.array(embedding_list)
                    elif isinstance(embedding, list):
                        embedding_vector = np.array(embedding)
                    else:
                        embedding_vector = np.array(embedding)
                    
                    dot_product = np.dot(question_vector, embedding_vector)
                    norm1 = np.linalg.norm(question_vector)
                    norm2 = np.linalg.norm(embedding_vector)
                    
                    if norm1 > 0 and norm2 > 0:
                        similarity = dot_product / (norm1 * norm2)
                        similar_chunks.append((similarity, doc_name, chunk_text))
                
                except Exception as calc_error:
                    continue
            
            similar_chunks.sort(reverse=True)
            return similar_chunks[:3]
    
    except Exception as e:
        print(f"Error finding chunks: {e}")
        return []

def general_rag_answer(user_question):
    """General RAG function (from working rag_system.py)"""
    print("🔍 Searching your documents...")
    
    chunks = find_similar_chunks_general(user_question)
    
    if not chunks:
        return "Sorry, couldn't find relevant information in your documents."
    
    context = "Relevant information from your documents:\n\n"
    sources = []
    
    for similarity, doc_name, chunk_text in chunks:
        context += f"From {doc_name}:\n{chunk_text}\n\n"
        if doc_name not in sources:
            sources.append(doc_name)
    
    prompt = f"""Answer the question based only on this information:

{context}

Question: {user_question}
Answer:"""
    
    print("🤖 Getting AI response...")
    
    try:
        response = requests.post(
            f"{OLLAMA_URL}/api/generate",
            json={"model": MODEL_NAME, "prompt": prompt, "stream": False},
            timeout=30
        )
        
        if response.status_code == 200:
            answer = response.json().get('response', 'No answer received')
        else:
            answer = f"AI Error: {response.status_code}"
    
    except Exception as e:
        answer = f"Error asking AI: {e}"
    
    answer += f"\n\n📚 Sources: {', '.join(sources)}"
    return answer

# ----------------------------
# MAIN INTERFACE
# ----------------------------

def main():
    print("🧪 AUTOMATED TEST CASE GENERATION SYSTEM")
    print("=" * 50)
    print("SPECIALIZED FOR: Test Case & Scenario Generation")
    print("INCLUDES: General RAG functionality from rag_system.py")
    print("=" * 50)
    
    # Check database
    try:
        with engine.connect() as conn:
            conn.execute(text(f"SET search_path TO {DB_SCHEMA}, public"))
            result = conn.execute(text("SELECT COUNT(*) FROM vector_documents"))
            total = result.fetchone()[0]
            
            excel_result = conn.execute(text("SELECT COUNT(*) FROM vector_documents WHERE file_type = 'excel'"))
            excel_count = excel_result.fetchone()[0]
            
            word_result = conn.execute(text("SELECT COUNT(*) FROM vector_documents WHERE file_type = 'word'"))
            word_count = word_result.fetchone()[0]
            
            print(f"📊 Database Status:")
            print(f"   Total chunks: {total}")
            print(f"   Excel chunks (test templates): {excel_count}")
            print(f"   Word chunks (workflow docs): {word_count}")
            print()
            
    except Exception as e:
        print(f"❌ Database error: {e}")
        return
    
    while True:
        print("\n" + "="*50)
        print("CHOOSE MODE:")
        print("1. 🧪 Generate Test Cases (Excel Export)")
        print("2. 💬 General Questions (RAG Mode)")
        print("3. 🚪 Quit")
        
        choice = input("\nEnter choice (1/2/3): ").strip()
        
        if choice == '3' or choice.lower() in ['quit', 'exit']:
            print("👋 Goodbye!")
            break
        
        elif choice == '1':
            # TEST CASE GENERATION MODE
            user_request = input("\n🎯 Describe what test cases you need: ").strip()
            
            if not user_request:
                continue
            
            if not is_test_case_request(user_request):
                print(refuse_non_test_queries(user_request))
                continue
            
            print(f"\n🚀 Processing request: {user_request}")
            print("=" * 50)
            
            # Generate test cases using RAG context
            ai_response = generate_test_cases_with_rag(user_request)
            
            print("\n📝 Generated Response:")
            print("-" * 20)
            print(ai_response[:500] + "..." if len(ai_response) > 500 else ai_response)
            
            # Export to Excel
            print("\n💾 Creating Excel file...")
            try:
                filepath, count = create_excel_from_response(ai_response, "automated_test_cases")
                
                print(f"✅ TEST CASES GENERATED SUCCESSFULLY!")
                print(f"📁 Excel file saved: {os.path.basename(filepath)}")
                print(f"📊 Total test cases: {count}")
                print(f"🗂️  Saved to folder: test_cases/")
                print(f"🔍 Full path: {os.path.abspath(filepath)}")
                
                print(f"\n📋 HOW TO ACCESS YOUR TEST CASES:")
                print(f"1. Open folder: test_cases/")
                print(f"2. Open file: {os.path.basename(filepath)}")
                print(f"3. Review generated test scenarios and cases")
                print(f"4. Modify and execute as needed")
                
            except Exception as e:
                print(f"❌ Error creating Excel: {e}")
        
        elif choice == '2':
            # GENERAL RAG MODE
            user_question = input("\n💬 Ask any question about your documents: ").strip()
            
            if not user_question:
                continue
            
            print(f"\n🔍 Processing: {user_question}")
            answer = general_rag_answer(user_question)
            print(f"\n💡 Answer:\n{answer}\n")
        
        else:
            print("❌ Invalid choice. Please enter 1, 2, or 3.")
        
        print("\n" + "="*50)

if __name__ == "__main__":
    main()
